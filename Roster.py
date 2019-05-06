# Python program to read and write an excel file Roster
# Han Phuoc Le 05/05/2019
# import openpyxl module
import openpyxl
import sys

# location of the file, please modify this path to your correct file location
path = "/Users/HanLe/Desktop/Jones_2019.xlsx"

# workbook object is created for reading
wb_obj = openpyxl.load_workbook(path, data_only = True)
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row
#workbook object is created for writing
wb_obj_towrite = openpyxl.load_workbook(path, data_only = False)
sheet_obj1 = wb_obj_towrite.active
max_row1 = sheet_obj1.max_row


print ("Please select an option: ")
print ("1. Display and/or update student grade")
print ("2. Delete a student record")
print ("3. Exit the program and call it a day")

# to test if user input is valid or not, if not then keep asking until valid
def validate_choice(choice):
    if choice == "1" or choice == "2" or choice == "3":
        return True
    return False

# delete student record function by their ID
def DeleteStudent(student_ID):
    selectedStudent = "Student_" + str(student_ID)
    selectedStudentSheet = wb_obj_towrite.get_sheet_by_name(selectedStudent)
    RosterSheet = wb_obj_towrite.get_sheet_by_name("Roster")
    for row in range(1, max_row1 + 1):
        if RosterSheet.cell(row, 1).value == student_ID:
            RosterSheet.delete_rows(row)
            wb_obj_towrite.remove_sheet(selectedStudentSheet)
            wb_obj_towrite.save(path)

# update student record function by their ID and assignment number
def UpdateStudentGrade(student_ID):
    studentDetailforUpdate = "Student_" + str(student_ID)
    studentDetailSheetforUpdate = wb_obj_towrite.get_sheet_by_name(studentDetailforUpdate)
    updateGradeChoice = int(input("Which assignment number do you want to update the grade? "))
    updatedGradeToSave = int(input("New grade you want to update: "))
    for rowInStudentSheet in range(6, 16):
        if studentDetailSheetforUpdate.cell(rowInStudentSheet, 1).value == updateGradeChoice:
            studentDetailSheetforUpdate.cell(rowInStudentSheet, 2).value = updatedGradeToSave
            wb_obj_towrite.save(path)
    return

# while loop to check for valid input from user
while True:
    try:
        choice = input("Which option would you like (1, 2, or 3)? ")
        if validate_choice(choice): break
        else:
            print ("Error: Invalid choice. Please try again.")

    except ValueError:
        print ("Error: Invalid choice. Please try again.")

# if user wants to see student grade and/or update them
if choice == "1":
    while True:
        try:
            findStudentChoice = input("Would you like to get student information by ID or Name? Type 1 for ID and 2 for Name: ")
            # look for student by ID
            if findStudentChoice == "1":
                student_ID = int(input("Please enter student ID: "))
                testStringIfStudentFound = ""
                # since openpyxl will lose the formula when I try to write the updated grade to the file and when call to display the grade again it will return as None so I have to calculate the average grade in Python with the help of sum of total grades divides by number of assignments
                countNumberofGrades = 0
                sumofAllGrades = 0
                # loop through each row to get the row with the selected student ID
                for row in range(1, max_row + 1):
                    if sheet_obj.cell(row, 1).value == student_ID:
                        studentDetail = "Student_" + str(student_ID)
                        studentDetailSheet = wb_obj.get_sheet_by_name(studentDetail)
                        # loop through assignment 1 to 10 to get total sum of grades and number of assignments
                        for roww in range(6, 16):
                            sumofAllGrades = sumofAllGrades + studentDetailSheet.cell(roww, 2).value
                            countNumberofGrades = countNumberofGrades + 1
                        # this string is to test if there is any existing student with selected ID
                        testStringIfStudentFound = testStringIfStudentFound + "studentFound"
                        # print out their info include ID, name, and average grade
                        print("ID: " + str(sheet_obj.cell(row, 1).value) + "|| First name: " + sheet_obj.cell(row, 2).value + "|| Last name: " + sheet_obj.cell(row, 3).value + "|| Grade: " + str(sumofAllGrades/countNumberofGrades))
                        # print out their detail grades
                        print("Assignment No.   Grade")
                        for roww in range(6, 16):
                            print(str(studentDetailSheet.cell(roww, 1).value).rjust(10) + str(studentDetailSheet.cell(roww, 2).value).rjust(10))
            # ask user if they want to update the student grade or just exit the program
                UpdateChoice = input("Do you want to update this student's grandes? Y for Yes and N for No ")
                if UpdateChoice == "Y" or UpdateChoice == "y" or UpdateChoice == "yes" or UpdateChoice == "Yes":
                    # if yes then call this update function
                    UpdateStudentGrade(student_ID)
                elif UpdateChoice == "N" or UpdateChoice == "n" or UpdateChoice == "no":
                    sys.exit()
                # if there is no student found with entered ID then show the message and exit program
                if testStringIfStudentFound == "":
                    print("No student found with the ID you entered. Program terminated.")
                
                break

            # look for student by their name
            elif findStudentChoice == "2":
                firstname = input("Enter student first name: ")
                lastname  = input("Enter student last name: ")
                testStringIfStudentFound = ""
                countNumberofGrades = 0
                sumofAllGrades = 0
                for rowww in range(1, max_row + 1):
                    if sheet_obj.cell(rowww, 2).value == firstname and sheet_obj.cell(rowww, 3).value == lastname:
                        studentDetail = "Student_" + str(sheet_obj.cell(rowww, 1).value)
                        studentDetailSheet = wb_obj.get_sheet_by_name(studentDetail)
                        sumofAllGrades = sumofAllGrades + studentDetailSheet.cell(roww, 2).value
                        countNumberofGrades = countNumberofGrades + 1
                        testStringIfStudentFound += "StudentFound"
                        print("ID: " + str(sheet_obj.cell(rowww, 1).value) + "|| First name: " + sheet_obj.cell(rowww, 2).value + "|| Last name: " + sheet_obj.cell(rowww, 3).value + "|| Grade: " + str(sheet_obj.cell(rowww, 4).value))
                        
                        print("Assignment No.   Grade")
                        for rowwww in range(6, 16):
                            print(str(studentDetailSheet.cell(rowwww, 1).value).rjust(10) + str(studentDetailSheet.cell(rowwww, 2).value).rjust(10))
                UpdateChoice = input("Do you want to update this student's grandes? Y for Yes and N for No ")
                if UpdateChoice == "Y" or UpdateChoice == "y" or UpdateChoice == "yes" or UpdateChoice == "Yes":
                    UpdateStudentGrade(student_ID)
                else:
                    sys.exit()
                if testStringIfStudentFound == "":
                    print("No student found with first and last name you entered. Program terminated.")
                break

            else:
                print("Invalid choice. Please try again.")
        except ValueError:
            print("Invalid choice. Please try again.")

# if user wants to delete a student record by their ID
elif choice == "2":
    deleteChoice = int(input("Please enter ID of student you want to delete: "))
    # call the delete function
    DeleteStudent(deleteChoice)
    print("Student with ID: " + str(deleteChoice) + " is successfully deleted.")

# exit the program
elif choice == "3":
    sys.exit()

